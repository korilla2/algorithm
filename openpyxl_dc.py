from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

# 새로운 엑셀 파일은 수동으로 만든 뒤 테스트
# Sheet1, Sheet2 두개의 시트를 생성한 뒤 작업

# 불러올 엑셀 파일의 Sheet1
new_ws_1 = load_workbook('test.xlsx')['Sheet1']
# 불러올 엑셀 파일의 Sheet2
new_ws_2 = load_workbook('test.xlsx')['Sheet2']

# 다른 엑셀파일의 워크시트를 추가하기 위한 작업
# 서로 다른 엑셀의 시트를 추가 할 수 없기때문에 인식하기 위해 지정해줌
new_ws_1._parent = wb
new_ws_2._parent = wb

# 본파일에 불러온 엑셀파일의 시트들을 추가
# 17번 18번 코드 없이하면 여기서 오류발생
wb._add_sheet(new_ws_1)
wb._add_sheet(new_ws_2)

wb.save("본파일.xlsx")
